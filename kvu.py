# -*- coding: utf-8 -*-

import sys
import getopt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill
from openpyxl.styles import colors

in_old_filename = u"//Users/alekseydorofeev/kvu/Форма реестра.xlsx"
out_old_filename = u"//Users/alekseydorofeev/kvu/Форма реестра_старая.xlsx"
in_new_filename = u"//Users/alekseydorofeev/kvu/Форма реестра1.xlsx"
out_new_filename = u"//Users/alekseydorofeev/kvu/Форма реестра1_новая.xlsx"
sheet_name = u"Таблица_ЗЛ"
new_sheet_name = u"Таблица_ЗЛ_изменения"

rft = Font(color=colors.RED)
bft = Font(color=colors.BLUE)

def found_changed_cols(row1, row2):
    cols = []
    i = 0
    for v in row1:
        if v.value != row2[i].value:
            cols.append(i)
        i = i+1
    return cols

def compare_files(filename1, filename2, new_filename):
    print "compare_files(%s, %s, %s)" % (filename1, filename2, new_filename)
    
    wb1 = load_workbook(filename=filename1)
    ws1 = wb1[sheet_name]

    wb2 = load_workbook(filename=filename2)
    ws2 = wb2[sheet_name]

    i1 = 0
    for row1 in ws1.rows:
        i1 = i1+1
        if i1 == 1:
            continue

        bill_num = row1[1].value
        name = row1[4].value
        typ = row1[5].value
        passport = row1[6].value
        uaddr = row1[7].value
        paddr = row1[8].value
        ao = row1[9].value
        ag = row1[10].value
    
        is_found = False
        is_changed = False
        changed_cols = []
        i2 = 0
        for row2 in ws2.rows:
            i2 = i2+1
            if i2 == 1:
                continue
            if bill_num == row2[1].value and typ == row2[5].value:
                if typ == u"ОС":
                    is_found = True
                    is_changed = (name!=row2[4].value or ao!=row2[9].value or ag!=row2[10].value)
                elif typ == u"ФЛ":
                    if name == row2[4].value:
                        is_found = True
                        is_changed = (passport!=row2[6].value or uaddr!=row2[7].value or paddr!=row2[8].value or ao!=row2[9].value or ag!=row2[10].value)
                elif typ == u"ЮЛ":
                    is_found = True
                    is_changed = (name!=row2[4].value or uaddr!=row2[7].value or paddr!=row2[8].value or ao!=row2[9].value or ag!=row2[10].value)
                if is_found:
                    break

        if is_found:
            if is_changed:
                print 'row #%d(%d): *' % (i1,i2)
                row1[1].font = rft
                row2[1].font = rft
                cols = found_changed_cols(row1, row2)
                for c in cols:
                    row1[c].font = rft
                    row2[c].font = rft
            else:
                print 'row #%d(%d): v' % (i1,i2)
        else:
            print 'row #%d: -' % (i1)
            for c in row1:
                c.font = bft

    wb1.save(new_filename)

def usage(argv):
    print "Usage: " + argv[0] + " -o old_registry_filename -n new_registry_filename"
    print " or"
    print "Usage: " + argv[0] + " --old=old_registry_filename --new=new_registry_filename"

def main(argv=None):
    if argv is None:
        argv = sys.argv
    # Разбираем аргументы командной строки
    try:
        opts, args = getopt.getopt(argv[1:], "ho:n:", ["help", "old=", "new="])
    except getopt.error, msg:
        print msg
        print "для справки используйте --help"
        sys.exit(2)
    #print >> sys.stderr, opts

    # Анализируем опции
    for o, a in opts:
        if o in ("-h", "--help"):
            usage(argv)
            sys.exit(0)
        elif o in ("-o", "--old"):
            print "old registry:" + a
            in_old_filename = a
            out_old_filename = "diff_old_registry.xlsx"
        elif o in ("-n", "--new"):
            print "new registry:" + a
            in_new_filename = a
            out_new_filename = "diff_new_registry.xlsx"
    
    # Анализируем аргументы
    #for arg in args:
    #    print "arg: " + arg

    compare_files(in_old_filename, in_new_filename, out_old_filename)
    compare_files(in_new_filename, in_old_filename, out_new_filename)

if __name__ == "__main__":
    main()

